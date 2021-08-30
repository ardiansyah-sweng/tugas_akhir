from nltk.corpus import stopwords
from flask import Flask, jsonify, request, redirect, url_for
from flask_restful import Resource, Api
from flask_cors import CORS
import pandas as pd
from openpyxl import load_workbook
import re
from nltk.tokenize import word_tokenize
from nltk.probability import FreqDist
import string
import nltk
import nltk
from Sastrawi.Stemmer.StemmerFactory import StemmerFactory
from sklearn.feature_extraction.text import CountVectorizer
nltk.download('stopwords')
nltk.download('punkt')


app = Flask(__name__)

api = Api(app)


class TrendResource(Resource):
    def get(self):
        wb = load_workbook(
            filename='data-genap1819.xlsx')
        sheet_ranges = wb['Sheet1']
        df = pd.DataFrame(sheet_ranges.values)
        da = df.dropna()
        da = da.iloc[1:]
        da.columns = ['judul', 'topik']
        popPS = da['topik'].value_counts()['Pengembangan Software']
        popFD = da['topik'].value_counts()['Forensik Digital']
        popKK = da['topik'].value_counts()['Keamanan Komputer']
        popSPK = da['topik'].value_counts()['Sistem Pendukung Keputusan']
        popJK = da['topik'].value_counts()['Jaringan Komputer']
        popML = da['topik'].value_counts()['Machine Learning']
        popDM = da['topik'].value_counts()['Data Mining']
        popAP = da['topik'].value_counts()['Algoritma Pencarian']
        popMP = da['topik'].value_counts()['Media Pembelajaran']
        popPC = da['topik'].value_counts()['Pengolahan Citra']
        popPBA = da['topik'].value_counts()['Pengolahan Bahasa Alami']
        popIMK = da['topik'].value_counts()['Interaksi Manusia dan Komputer']
        popSP = da['topik'].value_counts()['Sistem Pakar']
        popM = da['topik'].value_counts()['Multimedia']
        popK = da['topik'].value_counts()['Kriptografi']
        popG = da['topik'].value_counts()['Game']
        popWI = da['topik'].value_counts()['Web Indexing']
        array = [popPS, popFD, popKK, popSPK, popJK, popML, popDM, popAP,
                 popMP, popPC, popPBA, popIMK, popSP, popM, popK, popG, popWI]
        intps = int(array[0])
        intfd = int(array[1])
        intkk = int(array[2])
        intspk = int(array[3])
        intjk = int(array[4])
        intml = int(array[5])
        intdm = int(array[6])
        intap = int(array[7])
        intmp = int(array[8])
        intpc = int(array[9])
        intpba = int(array[10])
        intimk = int(array[11])
        intsp = int(array[12])
        intm = int(array[13])
        intk = int(array[14])
        intg = int(array[15])
        intwi = int(array[16])

        da['judul'] = da['judul'].str.lower()

        def remove(text):
            text = text.replace('\\t', " ").replace(
                '\\u', " ").replace('\\', "")
            text = text.encode('ascii', 'replace').decode('ascii')
            text = ' '.join(
                re.sub("([@#][A-Za-z0-9]+)|(\w+:\/\/\S+)", " ", text).split())
            return text.replace("http://", " ").replace("https://", " ")
        da['judul'] = da['judul'].apply(remove)

        def remove_number(text):
            return re.sub(r"\d+", "", text)
        da['judul'] = da['judul'].apply(remove_number)

        def remove_punctuaction(text):
            return text.translate(str.maketrans("", "", string.punctuation))
        da['judul'] = da['judul'].apply(remove_punctuaction)

        def remove_whitespace_LT(text):
            return text.strip()
        da['judul'] = da['judul'].apply(remove_whitespace_LT)

        def remove_whitespace_multiple(text):
            return re.sub('\s+', ' ', text)
        da['judul'] = da['judul'].apply(remove_whitespace_multiple)

        def remove_singl_char(text):
            return re.sub(r"\b[a-zA-Z]\b", "", text)
        da['judul'] = da['judul'].apply(remove_singl_char)

        def word_tokenize_wrapper(text):
            return word_tokenize(text)
        da['judul'] = da['judul'].apply(word_tokenize_wrapper)

        def stopword_removal(judul):
            filtering = stopwords.words('indonesian', 'english')
            x = []
            data = []

            def myfunc(x):
                if x in filtering:
                    return False
                else:
                    return True
            fit = filter(myfunc, judul)
            for x in fit:
                data.append(x)
            return data
        da['judul'] = da['judul'].apply(stopword_removal)

        factory = StemmerFactory()
        stemmer = factory.create_stemmer()

        def stemmed_wrapper(term):
            return stemmer.stem(term)

        term_dict = {}

        for document in da['judul']:
            for term in document:
                if term not in term_dict:
                    term_dict[term] = ' '

        for term in term_dict:
            term_dict[term] = stemmed_wrapper(term)

        def get_stemmed_term(document):
            return [term_dict[term] for term in document]
        da['judul'] = da['judul'].apply(get_stemmed_term)

        da[['judul', 'topik']].to_excel("databersih.xlsx")

        databersih = pd.read_excel(
            'databersih.xlsx')
        databersih.shape
        databersih

        #######################################
        ps = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Pengembangan Software']
        vec_PS = CountVectorizer()
        x_ps = vec_PS.fit_transform(ps)
        word_list_PS = vec_PS.get_feature_names()
        count_list_PS = x_ps.toarray().sum(axis=0)
        freq_PS = dict(zip(word_list_PS, count_list_PS))
        prob_PS = []
        for word, count in zip(word_list_PS, count_list_PS):
            prob_PS.append(count/len(word_list_PS))

        i = 0
        prob_PS_trend = 1
        for i in range(len(prob_PS)):
            prob_PS_trend = prob_PS_trend * prob_PS[i]
        ##############################################
        fd = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Forensik Digital']
        vec_FD = CountVectorizer()
        x_fd = vec_FD.fit_transform(fd)
        word_list_FD = vec_FD.get_feature_names()
        count_list_FD = x_fd.toarray().sum(axis=0)
        freq_FD = dict(zip(word_list_FD, count_list_FD))
        prob_FD = []
        for word, count in zip(word_list_FD, count_list_FD):
            prob_FD.append(count/len(word_list_FD))

        i = 0
        prob_FD_trend = 1
        for i in range(len(prob_FD)):
            prob_FD_trend = prob_FD_trend * prob_FD[i]
        ##############################################
        kk = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Keamanan Komputer']
        vec_KK = CountVectorizer()
        x_kk = vec_KK.fit_transform(kk)
        word_list_KK = vec_KK.get_feature_names()
        count_list_KK = x_kk.toarray().sum(axis=0)
        freq_KK = dict(zip(word_list_KK, count_list_KK))
        prob_KK = []
        for word, count in zip(word_list_KK, count_list_KK):
            prob_KK.append(count/len(word_list_KK))

        i = 0
        prob_KK_trend = 1
        for i in range(len(prob_KK)):
            prob_KK_trend = prob_KK_trend * prob_KK[i]
        ##############################################
        spk = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Sistem Pendukung Keputusan']
        vec_SPK = CountVectorizer()
        x_spk = vec_SPK.fit_transform(spk)
        word_list_SPK = vec_SPK.get_feature_names()
        count_list_SPK = x_spk.toarray().sum(axis=0)
        freq_SPK = dict(zip(word_list_SPK, count_list_SPK))
        prob_SPK = []
        for word, count in zip(word_list_SPK, count_list_SPK):
            prob_SPK.append(count/len(word_list_SPK))

        i = 0
        prob_SPK_trend = 1
        for i in range(len(prob_SPK)):
            prob_SPK_trend = prob_SPK_trend * prob_SPK[i]
        ##############################################
        jk = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Jaringan Komputer']
        vec_JK = CountVectorizer()
        x_jk = vec_JK.fit_transform(jk)
        word_list_JK = vec_JK.get_feature_names()
        count_list_JK = x_jk.toarray().sum(axis=0)
        freq_JK = dict(zip(word_list_JK, count_list_JK))
        prob_JK = []
        for word, count in zip(word_list_JK, count_list_JK):
            prob_JK.append(count/len(word_list_JK))

        i = 0
        prob_JK_trend = 1
        for i in range(len(prob_JK)):
            prob_JK_trend = prob_JK_trend * prob_JK[i]
        ##############################################
        ml = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Machine Learning']
        vec_ML = CountVectorizer()
        x_ml = vec_ML.fit_transform(ml)
        word_list_ML = vec_ML.get_feature_names()
        count_list_ML = x_ml.toarray().sum(axis=0)
        freq_ML = dict(zip(word_list_ML, count_list_ML))
        prob_ML = []
        for word, count in zip(word_list_ML, count_list_ML):
            prob_ML.append(count/len(word_list_ML))

        i = 0
        prob_ML_trend = 1
        for i in range(len(prob_ML)):
            prob_ML_trend = prob_ML_trend * prob_ML[i]
        ##############################################
        dm = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Data Mining']
        vec_DM = CountVectorizer()
        x_dm = vec_DM.fit_transform(dm)
        word_list_DM = vec_DM.get_feature_names()
        count_list_DM = x_dm.toarray().sum(axis=0)
        freq_DM = dict(zip(word_list_DM, count_list_DM))
        prob_DM = []
        for word, count in zip(word_list_DM, count_list_DM):
            prob_DM.append(count/len(word_list_DM))

        i = 0
        prob_DM_trend = 1
        for i in range(len(prob_DM)):
            prob_DM_trend = prob_DM_trend * prob_DM[i]
        ##############################################
        ap = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Algoritma Pencarian']
        vec_AP = CountVectorizer()
        x_ap = vec_AP.fit_transform(ap)
        word_list_AP = vec_AP.get_feature_names()
        count_list_AP = x_ap.toarray().sum(axis=0)
        freq_AP = dict(zip(word_list_AP, count_list_AP))
        prob_AP = []
        for word, count in zip(word_list_AP, count_list_AP):
            prob_AP.append(count/len(word_list_AP))

        i = 0
        prob_AP_trend = 1
        for i in range(len(prob_AP)):
            prob_AP_trend = prob_AP_trend * prob_AP[i]
        ##############################################
        mp = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Media Pembelajaran']
        vec_MP = CountVectorizer()
        x_mp = vec_MP.fit_transform(mp)
        word_list_MP = vec_MP.get_feature_names()
        count_list_MP = x_mp.toarray().sum(axis=0)
        freq_MP = dict(zip(word_list_MP, count_list_MP))
        prob_MP = []
        for word, count in zip(word_list_MP, count_list_MP):
            prob_MP.append(count/len(word_list_MP))

        i = 0
        prob_MP_trend = 1
        for i in range(len(prob_MP)):
            prob_MP_trend = prob_MP_trend * prob_MP[i]
        ##############################################
        pc = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Pengolahan Citra']
        vec_PC = CountVectorizer()
        x_pc = vec_PC.fit_transform(pc)
        word_list_PC = vec_PC.get_feature_names()
        count_list_PC = x_pc.toarray().sum(axis=0)
        freq_PC = dict(zip(word_list_PC, count_list_PC))
        prob_PC = []
        for word, count in zip(word_list_PC, count_list_PC):
            prob_PC.append(count/len(word_list_PC))

        i = 0
        prob_PC_trend = 1
        for i in range(len(prob_PC)):
            prob_PC_trend = prob_PC_trend * prob_PC[i]
        ##############################################
        pba = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Pengolahan Bahasa Alami']
        vec_PBA = CountVectorizer()
        x_pba = vec_PBA.fit_transform(pba)
        word_list_PBA = vec_PBA.get_feature_names()
        count_list_PBA = x_pba.toarray().sum(axis=0)
        freq_PBA = dict(zip(word_list_PBA, count_list_PBA))
        prob_PBA = []
        for word, count in zip(word_list_PBA, count_list_PBA):
            prob_PBA.append(count/len(word_list_PBA))

        i = 0
        prob_PBA_trend = 1
        for i in range(len(prob_PBA)):
            prob_PBA_trend = prob_PBA_trend * prob_PBA[i]
        ##############################################
        imk = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Interaksi Manusia dan Komputer']
        vec_IMK = CountVectorizer()
        x_imk = vec_IMK.fit_transform(imk)
        word_list_IMK = vec_IMK.get_feature_names()
        count_list_IMK = x_imk.toarray().sum(axis=0)
        freq_IMK = dict(zip(word_list_IMK, count_list_IMK))
        prob_IMK = []
        for word, count in zip(word_list_IMK, count_list_IMK):
            prob_IMK.append(count/len(word_list_IMK))

        i = 0
        prob_IMK_trend = 1
        for i in range(len(prob_IMK)):
            prob_IMK_trend = prob_IMK_trend * prob_IMK[i]
        ##############################################
        sp = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Sistem Pakar']
        vec_SP = CountVectorizer()
        x_sp = vec_SP.fit_transform(sp)
        word_list_SP = vec_SP.get_feature_names()
        count_list_SP = x_sp.toarray().sum(axis=0)
        freq_SP = dict(zip(word_list_SP, count_list_SP))
        prob_SP = []
        for word, count in zip(word_list_SP, count_list_SP):
            prob_SP.append(count/len(word_list_SP))

        i = 0
        prob_SP_trend = 1
        for i in range(len(prob_SP)):
            prob_SP_trend = prob_SP_trend * prob_SP[i]
        ##############################################
        m = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Multimedia']
        vec_M = CountVectorizer()
        x_m = vec_M.fit_transform(m)
        word_list_M = vec_M.get_feature_names()
        count_list_M = x_m.toarray().sum(axis=0)
        freq_M = dict(zip(word_list_M, count_list_M))
        prob_M = []
        for word, count in zip(word_list_M, count_list_M):
            prob_M.append(count/len(word_list_M))

        i = 0
        prob_M_trend = 1
        for i in range(len(prob_M)):
            prob_M_trend = prob_M_trend * prob_M[i]
        ##############################################
        k = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Kriptografi']
        vec_K = CountVectorizer()
        x_k = vec_K.fit_transform(k)
        word_list_K = vec_K.get_feature_names()
        count_list_K = x_k.toarray().sum(axis=0)
        freq_K = dict(zip(word_list_K, count_list_K))
        prob_K = []
        for word, count in zip(word_list_K, count_list_K):
            prob_K.append(count/len(word_list_K))

        i = 0
        prob_K_trend = 1
        for i in range(len(prob_K)):
            prob_K_trend = prob_K_trend * prob_K[i]
        ##############################################
        g = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Game']
        vec_G = CountVectorizer()
        x_g = vec_G.fit_transform(g)
        word_list_G = vec_G.get_feature_names()
        count_list_G = x_g.toarray().sum(axis=0)
        freq_G = dict(zip(word_list_G, count_list_G))
        prob_G = []
        for word, count in zip(word_list_G, count_list_G):
            prob_G.append(count/len(word_list_G))

        i = 0
        prob_G_trend = 1
        for i in range(len(prob_G)):
            prob_G_trend = prob_G_trend * prob_G[i]
        ##############################################
        wi = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Web indexing']
        vec_WI = CountVectorizer()
        x_wi = vec_WI.fit_transform(wi)
        word_list_WI = vec_WI.get_feature_names()
        count_list_WI = x_wi.toarray().sum(axis=0)
        freq_WI = dict(zip(word_list_WI, count_list_WI))
        prob_WI = []
        for word, count in zip(word_list_WI, count_list_WI):
            prob_WI.append(count/len(word_list_WI))

        i = 0
        prob_WI_trend = 1
        for i in range(len(prob_WI)):
            prob_WI_trend = prob_WI_trend * prob_WI[i]
        ##############################################

        response = [{"PS": prob_PS_trend, "FD": prob_FD_trend, "KK": prob_KK_trend, "SPK": prob_SPK_trend, "JK": prob_JK_trend, "ML": prob_ML_trend,
                     "DM": prob_DM_trend, "AP": prob_AP_trend, "MP": prob_MP_trend, "PC": prob_PC_trend, "PBA": prob_PBA_trend, "IMK": prob_IMK_trend,
                     "SP": prob_SP_trend, "M": prob_M_trend, "K": prob_K_trend, "G": prob_G_trend, "WI": prob_WI_trend,
                     "popPS": intps, "popFD": intfd, "popKK": intkk, "popSPK": intspk, "popJK": intjk, "popML": intml, "popDM": intdm,
                     "popAP": intap, "popMP": intmp, "popPC": intpc, "popPBA": intpba, "popIMK": intimk, "popSP": intsp, "popM": intm, "popK": intk,
                     "popG": intg, "popWI": intwi}]
        return response


api.add_resource(TrendResource, "/api", methods=["GET"])

api1 = Api(app)

rekomentopik = {}


class TrendResource1(Resource):
    def get(self):
        wb = load_workbook(
            filename='data-genap1819.xlsx')
        sheet_ranges = wb['Sheet1']
        df = pd.DataFrame(sheet_ranges.values)
        da = df.dropna()
        da = da.iloc[1:]
        da.columns = ['judul', 'topik']
        popPS = da['topik'].value_counts()['Pengembangan Software']
        popFD = da['topik'].value_counts()['Forensik Digital']
        popKK = da['topik'].value_counts()['Keamanan Komputer']
        popSPK = da['topik'].value_counts()['Sistem Pendukung Keputusan']
        popJK = da['topik'].value_counts()['Jaringan Komputer']
        popML = da['topik'].value_counts()['Machine Learning']
        popDM = da['topik'].value_counts()['Data Mining']
        popAP = da['topik'].value_counts()['Algoritma Pencarian']
        popMP = da['topik'].value_counts()['Media Pembelajaran']
        popPC = da['topik'].value_counts()['Pengolahan Citra']
        popPBA = da['topik'].value_counts()['Pengolahan Bahasa Alami']
        popIMK = da['topik'].value_counts()['Interaksi Manusia dan Komputer']
        popSP = da['topik'].value_counts()['Sistem Pakar']
        popM = da['topik'].value_counts()['Multimedia']
        popK = da['topik'].value_counts()['Kriptografi']
        popG = da['topik'].value_counts()['Game']
        popWI = da['topik'].value_counts()['Web Indexing']
        array = [popPS, popFD, popKK, popSPK, popJK, popML, popDM, popAP,
                 popMP, popPC, popPBA, popIMK, popSP, popM, popK, popG, popWI]
        intps = int(array[0])
        intfd = int(array[1])
        intkk = int(array[2])
        intspk = int(array[3])
        intjk = int(array[4])
        intml = int(array[5])
        intdm = int(array[6])
        intap = int(array[7])
        intmp = int(array[8])
        intpc = int(array[9])
        intpba = int(array[10])
        intimk = int(array[11])
        intsp = int(array[12])
        intm = int(array[13])
        intk = int(array[14])
        intg = int(array[15])
        intwi = int(array[16])

        da['judul'] = da['judul'].str.lower()

        def remove(text):
            text = text.replace('\\t', " ").replace(
                '\\u', " ").replace('\\', "")
            text = text.encode('ascii', 'replace').decode('ascii')
            text = ' '.join(
                re.sub("([@#][A-Za-z0-9]+)|(\w+:\/\/\S+)", " ", text).split())
            return text.replace("http://", " ").replace("https://", " ")
        da['judul'] = da['judul'].apply(remove)

        def remove_number(text):
            return re.sub(r"\d+", "", text)
        da['judul'] = da['judul'].apply(remove_number)

        def remove_punctuaction(text):
            return text.translate(str.maketrans("", "", string.punctuation))
        da['judul'] = da['judul'].apply(remove_punctuaction)

        def remove_whitespace_LT(text):
            return text.strip()
        da['judul'] = da['judul'].apply(remove_whitespace_LT)

        def remove_whitespace_multiple(text):
            return re.sub('\s+', ' ', text)
        da['judul'] = da['judul'].apply(remove_whitespace_multiple)

        def remove_singl_char(text):
            return re.sub(r"\b[a-zA-Z]\b", "", text)
        da['judul'] = da['judul'].apply(remove_singl_char)

        def word_tokenize_wrapper(text):
            return word_tokenize(text)
        da['judul'] = da['judul'].apply(word_tokenize_wrapper)

        def stopword_removal(judul):
            filtering = stopwords.words('indonesian', 'english')
            x = []
            data = []

            def myfunc(x):
                if x in filtering:
                    return False
                else:
                    return True
            fit = filter(myfunc, judul)
            for x in fit:
                data.append(x)
            return data
        da['judul'] = da['judul'].apply(stopword_removal)

        factory = StemmerFactory()
        stemmer = factory.create_stemmer()

        def stemmed_wrapper(term):
            return stemmer.stem(term)

        term_dict = {}

        for document in da['judul']:
            for term in document:
                if term not in term_dict:
                    term_dict[term] = ' '

        for term in term_dict:
            term_dict[term] = stemmed_wrapper(term)

        def get_stemmed_term(document):
            return [term_dict[term] for term in document]
        da['judul'] = da['judul'].apply(get_stemmed_term)

        da[['judul', 'topik']].to_excel("databersih.xlsx")
        databersih = pd.read_excel(
            'databersih.xlsx')
        databersih.shape
        databersih

        ps = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Pengembangan Software']
        vec_PS = CountVectorizer()
        x_ps = vec_PS.fit_transform(ps)
        word_list_PS = vec_PS.get_feature_names()
        count_list_PS = x_ps.toarray().sum(axis=0)
        freq_PS = dict(zip(word_list_PS, count_list_PS))
        ##############################################
        fd = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Forensik Digital']
        vec_FD = CountVectorizer()
        x_fd = vec_FD.fit_transform(fd)
        word_list_FD = vec_FD.get_feature_names()
        count_list_FD = x_fd.toarray().sum(axis=0)
        freq_FD = dict(zip(word_list_FD, count_list_FD))
        ##############################################
        kk = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Keamanan Komputer']
        vec_KK = CountVectorizer()
        x_kk = vec_KK.fit_transform(kk)
        word_list_KK = vec_KK.get_feature_names()
        count_list_KK = x_kk.toarray().sum(axis=0)
        freq_KK = dict(zip(word_list_KK, count_list_KK))
        ##############################################
        spk = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Sistem Pendukung Keputusan']
        vec_SPK = CountVectorizer()
        x_spk = vec_SPK.fit_transform(spk)
        word_list_SPK = vec_SPK.get_feature_names()
        count_list_SPK = x_spk.toarray().sum(axis=0)
        freq_SPK = dict(zip(word_list_SPK, count_list_SPK))
        ##############################################
        jk = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Jaringan Komputer']
        vec_JK = CountVectorizer()
        x_jk = vec_JK.fit_transform(jk)
        word_list_JK = vec_JK.get_feature_names()
        count_list_JK = x_jk.toarray().sum(axis=0)
        freq_JK = dict(zip(word_list_JK, count_list_JK))
        ##############################################
        ml = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Machine Learning']
        vec_ML = CountVectorizer()
        x_ml = vec_ML.fit_transform(ml)
        word_list_ML = vec_ML.get_feature_names()
        count_list_ML = x_ml.toarray().sum(axis=0)
        freq_ML = dict(zip(word_list_ML, count_list_ML))
        ##############################################
        dm = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Data Mining']
        vec_DM = CountVectorizer()
        x_dm = vec_DM.fit_transform(dm)
        word_list_DM = vec_DM.get_feature_names()
        count_list_DM = x_dm.toarray().sum(axis=0)
        freq_DM = dict(zip(word_list_DM, count_list_DM))
        ##############################################
        ap = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Algoritma Pencarian']
        vec_AP = CountVectorizer()
        x_ap = vec_AP.fit_transform(ap)
        word_list_AP = vec_AP.get_feature_names()
        count_list_AP = x_ap.toarray().sum(axis=0)
        freq_AP = dict(zip(word_list_AP, count_list_AP))
        ##############################################
        mp = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Media Pembelajaran']
        vec_MP = CountVectorizer()
        x_mp = vec_MP.fit_transform(mp)
        word_list_MP = vec_MP.get_feature_names()
        count_list_MP = x_mp.toarray().sum(axis=0)
        freq_MP = dict(zip(word_list_MP, count_list_MP))
        ##############################################
        pc = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Pengolahan Citra']
        vec_PC = CountVectorizer()
        x_pc = vec_PC.fit_transform(pc)
        word_list_PC = vec_PC.get_feature_names()
        count_list_PC = x_pc.toarray().sum(axis=0)
        freq_PC = dict(zip(word_list_PC, count_list_PC))
        ##############################################
        pba = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Pengolahan Bahasa Alami']
        vec_PBA = CountVectorizer()
        x_pba = vec_PBA.fit_transform(pba)
        word_list_PBA = vec_PBA.get_feature_names()
        count_list_PBA = x_pba.toarray().sum(axis=0)
        freq_PBA = dict(zip(word_list_PBA, count_list_PBA))
        ##############################################
        imk = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Interaksi Manusia dan Komputer']
        vec_IMK = CountVectorizer()
        x_imk = vec_IMK.fit_transform(imk)
        word_list_IMK = vec_IMK.get_feature_names()
        count_list_IMK = x_imk.toarray().sum(axis=0)
        freq_IMK = dict(zip(word_list_IMK, count_list_IMK))
        ##############################################
        sp = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Sistem Pakar']
        vec_SP = CountVectorizer()
        x_sp = vec_SP.fit_transform(sp)
        word_list_SP = vec_SP.get_feature_names()
        count_list_SP = x_sp.toarray().sum(axis=0)
        freq_SP = dict(zip(word_list_SP, count_list_SP))
        ##############################################
        m = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Multimedia']
        vec_M = CountVectorizer()
        x_m = vec_M.fit_transform(m)
        word_list_M = vec_M.get_feature_names()
        count_list_M = x_m.toarray().sum(axis=0)
        freq_M = dict(zip(word_list_M, count_list_M))
        ##############################################
        k = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Kriptografi']
        vec_K = CountVectorizer()
        x_k = vec_K.fit_transform(k)
        word_list_K = vec_K.get_feature_names()
        count_list_K = x_k.toarray().sum(axis=0)
        freq_K = dict(zip(word_list_K, count_list_K))
        ##############################################
        g = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Game']
        vec_G = CountVectorizer()
        x_g = vec_G.fit_transform(g)
        word_list_G = vec_G.get_feature_names()
        count_list_G = x_g.toarray().sum(axis=0)
        freq_G = dict(zip(word_list_G, count_list_G))
        ##############################################
        wi = [databersih['judul'] for index, databersih in databersih.iterrows(
        ) if databersih['topik'] == 'Web indexing']
        vec_WI = CountVectorizer()
        x_wi = vec_WI.fit_transform(wi)
        word_list_WI = vec_WI.get_feature_names()
        count_list_WI = x_wi.toarray().sum(axis=0)
        freq_WI = dict(zip(word_list_WI, count_list_WI))
        ##############################################

        docs = [databersih['judul']
                for index, databersih in databersih.iterrows()]
        vec = CountVectorizer()
        X = vec.fit_transform(docs)
        total_features = len(vec.get_feature_names())
        total_cnts_features_PS = count_list_PS.sum()
        total_cnts_features_FD = count_list_FD.sum()
        total_cnts_features_KK = count_list_KK.sum()
        total_cnts_features_SPK = count_list_SPK.sum()
        total_cnts_features_JK = count_list_JK.sum()
        total_cnts_features_ML = count_list_ML.sum()
        total_cnts_features_DM = count_list_DM.sum()
        total_cnts_features_AP = count_list_AP.sum()
        total_cnts_features_MP = count_list_MP.sum()
        total_cnts_features_PC = count_list_PC.sum()
        total_cnts_features_PBA = count_list_PBA.sum()
        total_cnts_features_IMK = count_list_IMK.sum()
        total_cnts_features_SP = count_list_SP.sum()
        total_cnts_features_M = count_list_M.sum()
        total_cnts_features_K = count_list_K.sum()
        total_cnts_features_G = count_list_G.sum()
        total_cnts_features_WI = count_list_WI.sum()

        masukanjudul = [rekomentopik["masukanjudul"]]

        tokenisasi_review = []

        for x in masukanjudul:
            lowerkal = x.lower()
            proses_token = word_tokenize(lowerkal)
            tokenisasi_review.append(proses_token)

        stopword1 = ['pada', 'yang', 'dalam', 'suatu',
                     'dengan', 'untuk', 'dan', 'di', 'sering']

        filtering1 = []
        for a in tokenisasi_review:
            for b in a:
                if b not in stopword1:
                    filtering1.append(b)

        factory = StemmerFactory()
        stemmer = factory.create_stemmer()

        datauji = []

        for c in filtering1:
            temp = stemmer.stem(c)
            datauji.append(temp)

        ####################################################
        prob_PS_with_ls = []
        for word in datauji:
            if word in freq_PS.keys():
                count = freq_PS[word]
            else:
                count = 0
            prob_PS_with_ls.append(
                (count + 1)/(total_cnts_features_PS + total_features))

        i = 0
        prob_PS_ = 1
        for i in range(len(prob_PS_with_ls)):
            prob_PS_ = prob_PS_ * prob_PS_with_ls[i]
        ######################################################
        prob_FD_with_ls = []
        for word in datauji:
            if word in freq_FD.keys():
                count = freq_FD[word]
            else:
                count = 0
            prob_FD_with_ls.append(
                (count + 1)/(total_cnts_features_FD + total_features))

        i = 0
        prob_FD_ = 1
        for i in range(len(prob_FD_with_ls)):
            prob_FD_ = prob_FD_ * prob_FD_with_ls[i]
        ##############################################################
        prob_KK_with_ls = []
        for word in datauji:
            if word in freq_KK.keys():
                count = freq_KK[word]
            else:
                count = 0
            prob_KK_with_ls.append(
                (count + 1)/(total_cnts_features_KK + total_features))
        dict(zip(datauji, prob_KK_with_ls))

        i = 0
        prob_KK_ = 1
        for i in range(len(prob_KK_with_ls)):
            prob_KK_ = prob_KK_ * prob_KK_with_ls[i]
        ###############################################################
        prob_SPK_with_ls = []
        for word in datauji:
            if word in freq_SPK.keys():
                count = freq_SPK[word]
            else:
                count = 0
            prob_SPK_with_ls.append(
                (count + 1)/(total_cnts_features_SPK + total_features))
        dict(zip(datauji, prob_SPK_with_ls))

        i = 0
        prob_SPK_ = 1
        for i in range(len(prob_SPK_with_ls)):
            prob_SPK_ = prob_SPK_ * prob_SPK_with_ls[i]
        ###################################################################
        prob_JK_with_ls = []
        for word in datauji:
            if word in freq_JK.keys():
                count = freq_JK[word]
            else:
                count = 0
            prob_JK_with_ls.append(
                (count + 1)/(total_cnts_features_JK + total_features))
        dict(zip(datauji, prob_JK_with_ls))

        i = 0
        prob_JK_ = 1
        for i in range(len(prob_JK_with_ls)):
            prob_JK_ = prob_JK_ * prob_JK_with_ls[i]
        ###################################################################
        prob_ML_with_ls = []
        for word in datauji:
            if word in freq_ML.keys():
                count = freq_ML[word]
            else:
                count = 0
            prob_ML_with_ls.append(
                (count + 1)/(total_cnts_features_ML + total_features))
        dict(zip(datauji, prob_ML_with_ls))

        i = 0
        prob_ML_ = 1
        for i in range(len(prob_ML_with_ls)):
            prob_ML_ = prob_ML_ * prob_ML_with_ls[i]
        ###################################################################
        prob_DM_with_ls = []
        for word in datauji:
            if word in freq_DM.keys():
                count = freq_DM[word]
            else:
                count = 0
            prob_DM_with_ls.append(
                (count + 1)/(total_cnts_features_DM + total_features))
        dict(zip(datauji, prob_DM_with_ls))

        i = 0
        prob_DM_ = 1
        for i in range(len(prob_DM_with_ls)):
            prob_DM_ = prob_DM_ * prob_DM_with_ls[i]
        ###################################################################
        prob_AP_with_ls = []
        for word in datauji:
            if word in freq_AP.keys():
                count = freq_AP[word]
            else:
                count = 0
            prob_AP_with_ls.append(
                (count + 1)/(total_cnts_features_AP + total_features))
        dict(zip(datauji, prob_AP_with_ls))

        i = 0
        prob_AP_ = 1
        for i in range(len(prob_AP_with_ls)):
            prob_AP_ = prob_AP_ * prob_AP_with_ls[i]
        ###################################################################
        prob_MP_with_ls = []
        for word in datauji:
            if word in freq_MP.keys():
                count = freq_MP[word]
            else:
                count = 0
            prob_MP_with_ls.append(
                (count + 1)/(total_cnts_features_MP + total_features))
        dict(zip(datauji, prob_MP_with_ls))

        i = 0
        prob_MP_ = 1
        for i in range(len(prob_MP_with_ls)):
            prob_MP_ = prob_MP_ * prob_MP_with_ls[i]
        ###################################################################
        prob_PC_with_ls = []
        for word in datauji:
            if word in freq_PC.keys():
                count = freq_PC[word]
            else:
                count = 0
            prob_PC_with_ls.append(
                (count + 1)/(total_cnts_features_PC + total_features))
        dict(zip(datauji, prob_PC_with_ls))

        i = 0
        prob_PC_ = 1
        for i in range(len(prob_PC_with_ls)):
            prob_PC_ = prob_PC_ * prob_PC_with_ls[i]
        ###################################################################
        prob_PBA_with_ls = []
        for word in datauji:
            if word in freq_PBA.keys():
                count = freq_PBA[word]
            else:
                count = 0
            prob_PBA_with_ls.append(
                (count + 1)/(total_cnts_features_PBA + total_features))
        dict(zip(datauji, prob_PBA_with_ls))

        i = 0
        prob_PBA_ = 1
        for i in range(len(prob_PBA_with_ls)):
            prob_PBA_ = prob_PBA_ * prob_PBA_with_ls[i]
        ###################################################################
        prob_IMK_with_ls = []
        for word in datauji:
            if word in freq_IMK.keys():
                count = freq_IMK[word]
            else:
                count = 0
            prob_IMK_with_ls.append(
                (count + 1)/(total_cnts_features_IMK + total_features))
        dict(zip(datauji, prob_IMK_with_ls))

        i = 0
        prob_IMK_ = 1
        for i in range(len(prob_IMK_with_ls)):
            prob_IMK_ = prob_IMK_ * prob_IMK_with_ls[i]
        ###################################################################
        prob_SP_with_ls = []
        for word in datauji:
            if word in freq_SP.keys():
                count = freq_SP[word]
            else:
                count = 0
            prob_SP_with_ls.append(
                (count + 1)/(total_cnts_features_SP + total_features))
        dict(zip(datauji, prob_SP_with_ls))

        i = 0
        prob_SP_ = 1
        for i in range(len(prob_SP_with_ls)):
            prob_SP_ = prob_SP_ * prob_SP_with_ls[i]
        ###################################################################
        prob_M_with_ls = []
        for word in datauji:
            if word in freq_M.keys():
                count = freq_M[word]
            else:
                count = 0
            prob_M_with_ls.append(
                (count + 1)/(total_cnts_features_M + total_features))
        dict(zip(datauji, prob_M_with_ls))

        i = 0
        prob_M_ = 1
        for i in range(len(prob_M_with_ls)):
            prob_M_ = prob_M_ * prob_M_with_ls[i]
        ###################################################################
        prob_K_with_ls = []
        for word in datauji:
            if word in freq_K.keys():
                count = freq_K[word]
            else:
                count = 0
            prob_K_with_ls.append(
                (count + 1)/(total_cnts_features_K + total_features))
        dict(zip(datauji, prob_K_with_ls))

        i = 0
        prob_K_ = 1
        for i in range(len(prob_K_with_ls)):
            prob_K_ = prob_K_ * prob_K_with_ls[i]
        ###################################################################
        prob_G_with_ls = []
        for word in datauji:
            if word in freq_G.keys():
                count = freq_G[word]
            else:
                count = 0
            prob_G_with_ls.append(
                (count + 1)/(total_cnts_features_G + total_features))
        dict(zip(datauji, prob_G_with_ls))

        i = 0
        prob_G_ = 1
        for i in range(len(prob_G_with_ls)):
            prob_G_ = prob_G_ * prob_G_with_ls[i]
        ###################################################################
        prob_WI_with_ls = []
        for word in datauji:
            if word in freq_WI.keys():
                count = freq_WI[word]
            else:
                count = 0
            prob_WI_with_ls.append(
                (count + 1)/(total_cnts_features_WI + total_features))
        dict(zip(datauji, prob_WI_with_ls))

        i = 0
        prob_WI_ = 1
        for i in range(len(prob_WI_with_ls)):
            prob_WI_ = prob_WI_ * prob_WI_with_ls[i]
        ###################################################################

        response = [{"value": prob_PS_, "value1": prob_FD_, "value2": prob_KK_, "value3": prob_SPK_, "value4": prob_JK_,
                     "value5": prob_ML_, "value6": prob_DM_, "value7": prob_AP_, "value8": prob_MP_, "value9": prob_PC_, "value10": prob_PBA_,
                     "value11": prob_IMK_, "value12": prob_SP_, "value13": prob_M_, "value14": prob_K_, "value15": prob_G_, "value16": prob_WI_}]
        return response, rekomentopik

    def post(self):

        hasil = {"status": "gagal"}
        try:
            data = request.json
            masukanjudul = data["masukanjudul"]
            rekomentopik["masukanjudul"] = masukanjudul
            hasil = {"status": "berhasil"}
        except Exception as e:
            print("Erorr : " + str(e))

        return jsonify(hasil)


api1.add_resource(TrendResource1, "/api1", methods=["GET", "POST"])

api2 = Api(app)

rekomentopik1 = {}


class TrendResource2(Resource):
    def get(self):

        return rekomentopik1

    def post(self):

        hasil1 = {"status": "gagal"}
        try:
            data = request.json
            nambahjudulexcel = data["nambahjudulexcel"]
            nambahtopikexcel = data["nambahtopikexcel"]
            rekomentopik1["nambahjudulexcel"] = nambahjudulexcel
            rekomentopik1["nambahtopikexcel"] = nambahtopikexcel
            wb = load_workbook(
                filename='data-genap1819.xlsx')
            sheet_ranges = wb['Sheet1']

            df = pd.DataFrame(sheet_ranges.values)
            da = df.dropna()
            da = da.iloc[1:]
            da.columns = ['judul', 'topik']
            judulBefore = []
            for dt in da['judul']:
                judulBefore.append(''.join(dt))
            topikBefore = []
            for dt1 in da['topik']:
                topikBefore.append(''.join(dt1))

            judulnew = [nambahjudulexcel]
            if nambahtopikexcel == '1':
                topiknew = ['Pengembangan Sistem']
            elif nambahtopikexcel == '2':
                topiknew = ['Forensik Digital']
            elif nambahtopikexcel == '3':
                topiknew = ['Keamanan Komputer']
            elif nambahtopikexcel == '4':
                topiknew = ['Sistem Pendukung Keputusan']
            elif nambahtopikexcel == '5':
                topiknew = ['Jaringan Komputer']
            elif nambahtopikexcel == '6':
                topiknew = ['Machine Learning']
            elif nambahtopikexcel == '7':
                topiknew = ['Data Mining']
            elif nambahtopikexcel == '8':
                topiknew = ['Media Pembelajaran']
            elif nambahtopikexcel == '9':
                topiknew = ['Pengolahan Citra']
            elif nambahtopikexcel == '10':
                topiknew = ['Pengolahan Bahasa Alami']
            elif nambahtopikexcel == '11':
                topiknew = ['Intraksi Manusia dan Komputer']
            elif nambahtopikexcel == '12':
                topiknew = ['Sistem Pakar']
            elif nambahtopikexcel == '13':
                topiknew = ['Multimedia']
            elif nambahtopikexcel == '14':
                topiknew = ['Kriotografi']
            elif nambahtopikexcel == '15':
                topiknew = ['Game']
            elif nambahtopikexcel == '16':
                topiknew = ['Web Indexing']
            else:
                topiknew = ['Tidak ada']

            data = [[judulnew, topiknew]]
            newDataframe = pd.DataFrame(data, columns=['judul', 'topik'])
            rowsJudulNew = []
            for d in newDataframe['judul']:
                rowsJudulNew.append(''.join(d))
            rowsTopikNew = []
            for d1 in newDataframe['topik']:
                rowsTopikNew.append(''.join(d1))
            for row in rowsJudulNew:
                judulBefore.append(''.join(row))
            for row1 in rowsTopikNew:
                topikBefore.append(''.join(row1))
            latestDataframe = pd.DataFrame(columns=['judul', 'topik'])
            latestDataframe['judul'] = judulBefore
            latestDataframe['topik'] = topikBefore
            latestDataframe.to_excel(
                'data-genap1819.xlsx', engine='xlsxwriter', index=False)
            hasil1 = {"status": "berhasil"}
        except Exception as e:
            print("Erorr : " + str(e))

        return jsonify(hasil1)


api2.add_resource(TrendResource2, "/api2", methods=["GET", "POST"])


# def casefolding(judul):
#     judul = judul.lower()
#     judul = judul.strip(" ")
#     judul = re.sub(r'[?|$|.|_:")(-+,]', '', judul)
#     return judul


if __name__ == "__main__":
    app.run(debug=True, port=5002)
